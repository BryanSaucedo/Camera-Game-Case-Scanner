import cv2
import openpyxl
import os
import requests

# Securely get OCR.Space API key
OCR_SPACE_API_KEY = os.getenv("OCR_SPACE_API_KEY", input("Enter OCR.Space API key: "))

# ----------- EXCEL SETUP -----------
excel_file = "GameLibrary.xlsx"
try:
    if not os.path.exists(excel_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Console", "Game Title"])
        wb.save(excel_file)
except PermissionError:
    print(f"Error: Cannot create {excel_file}. Please ensure the file is not open.")
    exit()
except Exception as e:
    print(f"Error creating Excel file: {e}")
    exit()

# ----------- OCR Function Using OCR.Space -----------
def extract_text_ocrspace(image_bytes, api_key=OCR_SPACE_API_KEY):
    print("📡 Sending image to OCR.Space...")
    url = "https://api.ocr.space/parse/image"
    payload = {
        'isOverlayRequired': False,
        'apikey': api_key,
        'language': 'eng'
    }
    files = {'filename': ('gamecase.jpg', image_bytes)}
    try:
        response = requests.post(url, files=files, data=payload)
        response.raise_for_status()
        result = response.json()
        if result.get("IsErroredOnProcessing"):
            error = result.get("ErrorMessage", ["Unknown error"])[0]
            if "API Key" in error:
                print("❌ Invalid OCR.Space API key. Please check your key.")
            elif "File size" in error:
                print("❌ Image file too large. Try a smaller resolution.")
            return ""
        return result['ParsedResults'][0]['ParsedText']
    except requests.RequestException as e:
        print("❌ OCR.Space request failed:", e)
        return ""

# ----------- MAIN SCAN LOOP -----------
while True:
    try:
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            raise Exception("Webcam not accessible.")
    except Exception as e:
        print(f"Error initializing webcam: {e}")
        exit()

    print("📸 Press SPACE to scan the game case, or ESC to exit.")

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Failed to capture image.")
            cap.release()
            cv2.destroyAllWindows()
            exit()

        cv2.imshow("Game Case Scanner", frame)
        key = cv2.waitKey(1)

        if key == 27:  # ESC
            print("Exiting scanner.")
            cap.release()
            cv2.destroyAllWindows()
            exit()
        elif key == 32:  # SPACE
            print("✅ Image captured.")
            image = frame
            break

    cap.release()
    cv2.destroyAllWindows()

    # ------ Encode full image (no preprocessing) ------
    success, encoded_image = cv2.imencode('.jpg', image)
    if not success:
        print("Error encoding image.")
        continue

    # ------ Single OCR call for full image ------
    full_text = extract_text_ocrspace(encoded_image.tobytes())
    if not full_text.strip():
        print("❌ No text detected.")
        retry = input("Retry scan? (y/n): ").lower()
        if retry == 'y':
            continue
        else:
            exit()

    # ------ Console detection (prioritize early lines) ------
    console_list = ["PS3", "PS4", "PS5", "Xbox 360", "Xbox One", "Xbox Series X", "Wii", "Switch", "GameCube"]
    console_detected = "Unknown Console"
    lines = [line.strip() for line in full_text.split("\n") if len(line.strip()) > 0]
    # Check first 2 lines for console (approximates upper Hawkins upper portion)
    for line in lines[:2]:
        for console in console_list:
            if console.lower() in line.lower():
                console_detected = console
                break
    # Fallback: Check all lines if not found in first 2
    if console_detected == "Unknown Console":
        for line in lines:
            for console in console_list:
                if console.lower() in line.lower():
                    console_detected = console
                    break
        if console_detected == "Unknown Console":
            print("❌ No console detected.")
            retry = input("Retry scan or manually enter console? (r/m/n): ").lower()
            if retry == 'r':
                continue
            elif retry == 'm':
                console_detected = input("Enter console: ")
            else:
                exit()

    # ------ Game title extraction from full text ------
    common_exclusions = ["edition", "version", "collector", "region"]
    game_title = "Unknown Title"
    filtered_lines = [line for line in lines if len(line.strip()) > 4]
    if filtered_lines:
        game_title = next((line for line in filtered_lines if not any(excl in line.lower() for excl in common_exclusions)), filtered_lines[0])
        if console_detected.lower() in game_title.lower() and len(filtered_lines) > 1:
            game_title = next((line for line in filtered_lines[1:] if not any(excl in line.lower() for excl in common_exclusions)), filtered_lines[1])

    print("\n📝 OCR Output:")
    print(full_text)
    print(f"\n🎮 Detected Console: {console_detected}")
    print(f"🎯 Detected Title: {game_title}")

    # ----------- OPTIONAL EDIT & RETRY -----------
    while True:
        retry = input("Retry scan or edit entry? (r/e/n): ").lower()
        if retry == 'r':
            break
        elif retry == 'e':
            console_detected = input("Enter console: ")
            game_title = input("Enter title: ")
            break
        elif retry == 'n':
            break
        else:
            print("Invalid input. Please enter 'r' (retry), 'e' (edit), or 'n' (continue).")
    if retry == 'r':
        continue

    # ----------- CONFIRM & SAVE -----------
    confirm = input("💾 Save this entry to Excel? (y/n): ").lower()
    if confirm != 'y':
        print("Entry not saved. Press SPACE to scan another game, or ESC to exit.")
        continue

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        existing = [(row[0], row[1]) for row in ws.iter_rows(min_row=2, values_only=True)]
        if (console_detected, game_title) in existing:
            print("⚠️ Entry already exists in Excel.")
            continue
        ws.append([console_detected, game_title])

        data = list(ws.iter_rows(min_row=2, values_only=True))
        data.sort(key=lambda x: (x[0].lower(), x[1].lower()))
        ws.delete_rows(2, ws.max_row)
        for row in data:
            ws.append(row)

        wb.save(excel_file)
        print(f"\n✅ Saved to {excel_file}: {console_detected} - {game_title}")
        print("Press SPACE to scan another game, or ESC to exit.")
    except PermissionError:
        print(f"Error: Cannot save to {excel_file}. Please close the file and try again.")
        continue
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        continue

#My API key(not given)
